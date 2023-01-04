import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

# List each company with respective product count

product_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value   # 4 is coloum number
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)


    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

#calculation total value of inventory per supplier
    # List each company with respective total inventory value
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price


# List products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)
# Write to SpreadSheet : calculate and write inventory value for each product into spreadsheet

    inventory_price.value = inventory * price


print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
inventory_file.save("inventory_with_total_value.xlsx")